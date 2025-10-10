"""
Excel document processor for spreadsheet data extraction.
"""

from typing import Optional, List, Dict, Any
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from app.core.processors.base_processor import BaseProcessor
from app.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelProcessor(BaseProcessor):
    """
    Processor for Excel documents using openpyxl.
    
    Supports:
    - .xlsx (Excel 2007+)
    - .xlsm (Excel with macros)
    - .xltx (Excel templates)
    - .xltm (Excel macro templates)
    
    For legacy .xls files, use xlrd library.
    """
    
    def __init__(self):
        """Initialize Excel processor."""
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xlsm', '.xltx', '.xltm']
        self.supported_mime_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel.sheet.macroEnabled.12',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.template',
            'application/vnd.ms-excel.template.macroEnabled.12'
        ]
    
    def can_process(self, file_path: str, mime_type: Optional[str] = None) -> bool:
        """
        Check if processor can handle Excel files.
        
        Args:
            file_path: Path to file
            mime_type: Optional MIME type
            
        Returns:
            True if file is Excel format
        """
        file = Path(file_path)
        
        # Check extension
        if file.suffix.lower() in self.supported_extensions:
            return True
        
        # Check MIME type
        if mime_type and mime_type in self.supported_mime_types:
            return True
        
        return False
    
    def extract_text(self, file_path: str) -> str:
        """
        Extract all text from Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Extracted text content from all sheets
        """
        try:
            text_parts = []
            
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            # Extract text from each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Add sheet name as header
                text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                
                # Extract text from cells
                sheet_text = self._extract_sheet_text(sheet)
                if sheet_text:
                    text_parts.append(sheet_text)
            
            workbook.close()
            
            # Combine all text
            full_text = '\n'.join(text_parts)
            
            # Clean text
            full_text = self.clean_text(full_text)
            
            logger.debug(f"Extracted {len(full_text)} characters from Excel")
            
            return full_text
            
        except Exception as e:
            logger.error(f"Error extracting text from Excel: {e}")
            raise
    
    def extract_tables(self, file_path: str) -> List[List[List[str]]]:
        """
        Extract tables from Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of tables (one per sheet)
        """
        try:
            all_tables = []
            
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            # Extract table from each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract table data
                table = self._extract_sheet_table(sheet)
                
                if table:
                    all_tables.append(table)
            
            workbook.close()
            
            logger.debug(f"Extracted {len(all_tables)} tables from Excel")
            
            return all_tables
            
        except Exception as e:
            logger.error(f"Error extracting tables from Excel: {e}")
            return []
    
    def extract_metadata(self, file_path: str) -> Dict[str, Any]:
        """
        Extract Excel file metadata.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with metadata
        """
        metadata = super().extract_metadata(file_path)
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            # Get workbook properties
            props = workbook.properties
            
            metadata.update({
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
                'title': props.title if props else None,
                'author': props.creator if props else None,
                'subject': props.subject if props else None,
                'created': props.created.isoformat() if props and props.created else None,
                'modified': props.modified.isoformat() if props and props.modified else None,
                'is_excel': True
            })
            
            # Get sheet statistics
            sheet_stats = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                stats = self._get_sheet_stats(sheet)
                sheet_stats.append({
                    'name': sheet_name,
                    **stats
                })
            
            metadata['sheets'] = sheet_stats
            
            workbook.close()
            
        except Exception as e:
            logger.warning(f"Error extracting Excel metadata: {e}")
        
        return metadata
    
    def _extract_sheet_text(self, sheet: Worksheet) -> str:
        """
        Extract text from a single sheet.
        
        Args:
            sheet: Worksheet object
            
        Returns:
            Text content
        """
        text_parts = []
        
        for row in sheet.iter_rows():
            row_values = []
            for cell in row:
                if cell.value is not None:
                    row_values.append(str(cell.value))
            
            if row_values:
                text_parts.append(' '.join(row_values))
        
        return '\n'.join(text_parts)
    
    def _extract_sheet_table(self, sheet: Worksheet) -> List[List[str]]:
        """
        Extract table data from a single sheet.
        
        Args:
            sheet: Worksheet object
            
        Returns:
            Table as list of rows
        """
        table = []
        
        for row in sheet.iter_rows():
            row_data = []
            has_data = False
            
            for cell in row:
                cell_value = cell.value
                if cell_value is not None:
                    row_data.append(str(cell_value))
                    has_data = True
                else:
                    row_data.append('')
            
            # Only include rows with at least one non-empty cell
            if has_data:
                table.append(row_data)
        
        return table
    
    def _get_sheet_stats(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        Get statistics for a sheet.
        
        Args:
            sheet: Worksheet object
            
        Returns:
            Statistics dictionary
        """
        stats = {
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'total_cells': sheet.max_row * sheet.max_column,
            'non_empty_cells': 0
        }
        
        # Count non-empty cells
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    stats['non_empty_cells'] += 1
        
        return stats
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get list of sheet names in Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet names
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            return sheet_names
            
        except Exception as e:
            logger.error(f"Error getting sheet names: {e}")
            return []
    
    def extract_sheet_by_name(self, file_path: str, sheet_name: str) -> List[List[str]]:
        """
        Extract data from a specific sheet by name.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to extract
            
        Returns:
            Sheet data as list of rows
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            
            sheet = workbook[sheet_name]
            table = self._extract_sheet_table(sheet)
            
            workbook.close()
            
            return table
            
        except Exception as e:
            logger.error(f"Error extracting sheet '{sheet_name}': {e}")
            raise
    
    def extract_sheet_by_index(self, file_path: str, sheet_index: int) -> List[List[str]]:
        """
        Extract data from a specific sheet by index.
        
        Args:
            file_path: Path to Excel file
            sheet_index: Index of sheet (0-based)
            
        Returns:
            Sheet data as list of rows
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            sheet_names = workbook.sheetnames
            if sheet_index < 0 or sheet_index >= len(sheet_names):
                raise ValueError(f"Sheet index {sheet_index} out of range")
            
            sheet_name = sheet_names[sheet_index]
            sheet = workbook[sheet_name]
            table = self._extract_sheet_table(sheet)
            
            workbook.close()
            
            return table
            
        except Exception as e:
            logger.error(f"Error extracting sheet at index {sheet_index}: {e}")
            raise
    
    def detect_header_row(self, sheet_data: List[List[str]]) -> int:
        """
        Detect the header row in sheet data.
        
        Args:
            sheet_data: Sheet data as list of rows
            
        Returns:
            Index of header row (0-based)
        """
        if not sheet_data:
            return 0
        
        # Simple heuristic: header row is the first non-empty row
        # with the most non-numeric values
        for i, row in enumerate(sheet_data):
            non_empty = [cell for cell in row if cell.strip()]
            if len(non_empty) >= 2:  # At least 2 non-empty cells
                # Check if mostly non-numeric
                non_numeric = sum(1 for cell in non_empty if not self._is_numeric(cell))
                if non_numeric / len(non_empty) > 0.5:
                    return i
        
        return 0
    
    def _is_numeric(self, value: str) -> bool:
        """Check if a string value is numeric."""
        try:
            float(value.replace(',', '').replace('$', ''))
            return True
        except (ValueError, AttributeError):
            return False
    
    def extract_with_headers(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, str]]:
        """
        Extract data with headers as dictionary keys.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (uses first sheet if None)
            
        Returns:
            List of row dictionaries with headers as keys
        """
        try:
            if sheet_name:
                sheet_data = self.extract_sheet_by_name(file_path, sheet_name)
            else:
                tables = self.extract_tables(file_path)
                sheet_data = tables[0] if tables else []
            
            if not sheet_data:
                return []
            
            # Detect header row
            header_index = self.detect_header_row(sheet_data)
            
            # Get headers and clean them
            headers = [str(cell).strip() for cell in sheet_data[header_index]]
            
            # Extract data rows
            data_rows = []
            for row in sheet_data[header_index + 1:]:
                if any(cell.strip() for cell in row):  # Skip empty rows
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = row[i].strip()
                        else:
                            row_dict[header] = ''
                    data_rows.append(row_dict)
            
            logger.debug(f"Extracted {len(data_rows)} rows with headers from Excel")
            
            return data_rows
            
        except Exception as e:
            logger.error(f"Error extracting with headers: {e}")
            return []
    
    def search_for_value(self, file_path: str, search_value: str) -> List[Dict[str, Any]]:
        """
        Search for a value in the Excel file.
        
        Args:
            file_path: Path to Excel file
            search_value: Value to search for
            
        Returns:
            List of matches with location information
        """
        try:
            matches = []
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        if cell.value and search_value.lower() in str(cell.value).lower():
                            matches.append({
                                'sheet': sheet_name,
                                'row': row_idx,
                                'column': col_idx,
                                'column_letter': get_column_letter(col_idx),
                                'cell_reference': f"{get_column_letter(col_idx)}{row_idx}",
                                'value': str(cell.value)
                            })
            
            workbook.close()
            
            return matches
            
        except Exception as e:
            logger.error(f"Error searching for value: {e}")
            return []
    
    def get_processor_info(self) -> Dict[str, Any]:
        """
        Get Excel processor information.
        
        Returns:
            Dictionary with processor details
        """
        info = super().get_processor_info()
        info.update({
            'can_extract_by_sheet': True,
            'can_detect_headers': True,
            'can_search': True,
            'can_get_statistics': True,
            'libraries': ['openpyxl']
        })
        return info


# Global Excel processor instance
_excel_processor: Optional[ExcelProcessor] = None


def get_excel_processor() -> ExcelProcessor:
    """
    Get or create Excel processor singleton.
    
    Returns:
        ExcelProcessor instance
    """
    global _excel_processor
    
    if _excel_processor is None:
        _excel_processor = ExcelProcessor()
    
    return _excel_processor